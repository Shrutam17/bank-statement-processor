from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.utils.logger import get_logger

logger = get_logger(__name__)

COLUMNS = [
    ("date", "Date"),
    ("description", "Description"),
    ("debit", "Debit"),
    ("credit", "Credit"),
    ("balance", "Balance"),
    ("category", "Category"),
    ("classification_confidence", "Confidence"),
]

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def export_to_excel(transactions, output_path, account_details=None):
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Account Summary"
    _write_summary_sheet(summary_sheet, account_details, transactions)

    txn_sheet = workbook.create_sheet("Transactions")
    _write_transactions_sheet(txn_sheet, transactions)

    workbook.save(output_path)
    logger.info(f"Exported {len(transactions)} transactions to Excel at {output_path}")
    return output_path


def _write_summary_sheet(sheet, account_details, transactions):
    account_details = account_details or {}
    rows = [
        ("Account Holder Name", account_details.get("account_holder_name") or "N/A"),
        ("Account Number", account_details.get("account_number") or "N/A"),
        ("IFSC Code", account_details.get("ifsc_code") or "N/A"),
        ("Branch", account_details.get("branch") or "N/A"),
        ("Total Transactions", len(transactions)),
        ("Total Debit", round(sum(t.get("debit", 0) for t in transactions), 2)),
        ("Total Credit", round(sum(t.get("credit", 0) for t in transactions), 2)),
    ]
    sheet["A1"] = "Field"
    sheet["B1"] = "Value"
    sheet["A1"].font = HEADER_FONT
    sheet["B1"].font = HEADER_FONT
    sheet["A1"].fill = HEADER_FILL
    sheet["B1"].fill = HEADER_FILL

    for row_index, (label, value) in enumerate(rows, start=2):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 30


def _write_transactions_sheet(sheet, transactions):
    for col_index, (_, header_label) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header_label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row_index, txn in enumerate(transactions, start=2):
        for col_index, (field, _) in enumerate(COLUMNS, start=1):
            sheet.cell(row=row_index, column=col_index, value=txn.get(field))

    for col_index, (field, header_label) in enumerate(COLUMNS, start=1):
        width = 18 if field == "description" else 14
        sheet.column_dimensions[get_column_letter(col_index)].width = width

    sheet.freeze_panes = "A2"
